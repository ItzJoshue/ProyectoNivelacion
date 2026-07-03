from openpyxl import load_workbook

from domain.entidades.estudiante import Estudiante
from domain.interfaces.importador import IImportadorEstudiantes
from factories.persona_factory import EstudianteFactory

COLUMNAS_ESPERADAS = ("cedula", "nombre", "apellido", "carrera", "email")


class ExcelImportadorEstudiantes(IImportadorEstudiantes):
    """Importa estudiantes desde un archivo Excel (.xlsx)."""

    def __init__(self, fabrica: EstudianteFactory | None = None) -> None:
        self._fabrica = fabrica or EstudianteFactory()

    def importar(self, ruta: str) -> list[Estudiante]:
        libro = load_workbook(ruta, read_only=True, data_only=True)
        hoja = libro.active
        filas = hoja.iter_rows(values_only=True)

        try:
            encabezados = next(filas)
        except StopIteration as exc:
            raise ValueError("El archivo Excel está vacío.") from exc

        indices = self._mapear_columnas(encabezados)
        estudiantes: list[Estudiante] = []

        for numero_fila, fila in enumerate(filas, start=2):
            if not fila or all(celda is None or str(celda).strip() == "" for celda in fila):
                continue

            datos = {
                columna: self._obtener_valor(fila, indices.get(columna))
                for columna in COLUMNAS_ESPERADAS
            }

            if not datos["cedula"] or not datos["nombre"] or not datos["apellido"]:
                raise ValueError(
                    f"Fila {numero_fila}: cédula, nombre y apellido son obligatorios."
                )

            estudiantes.append(self._fabrica.crear(datos))

        libro.close()
        return estudiantes

    @staticmethod
    def _mapear_columnas(encabezados: tuple) -> dict[str, int]:
        normalizados = {
            str(col).strip().lower(): indice
            for indice, col in enumerate(encabezados)
            if col is not None
        }

        faltantes = [col for col in ("cedula", "nombre", "apellido") if col not in normalizados]
        if faltantes:
            raise ValueError(
                f"Columnas obligatorias faltantes: {', '.join(faltantes)}. "
                f"Se esperan: {', '.join(COLUMNAS_ESPERADAS)}"
            )

        return normalizados

    @staticmethod
    def _obtener_valor(fila: tuple, indice: int | None) -> str:
        if indice is None or indice >= len(fila) or fila[indice] is None:
            return ""
        return str(fila[indice]).strip()
